# -*- coding: utf-8 -*-

from django_filters import rest_framework
from rest_framework.decorators import api_view, permission_classes
from rest_framework.generics import UpdateAPIView, DestroyAPIView
from rest_framework.parsers import JSONParser
from rest_framework.permissions import AllowAny, IsAuthenticated, IsAdminUser
from rest_framework.response import Response
from rest_framework import permissions
from rest_framework.views import APIView
from firebasesdkadmin import send_firebase_message, send_firebase_message1
from page.filters import *
from .models import *
from rest_framework_jwt.settings import api_settings
from page.pagination import LargeResultsSetPagination
from page.serializers import *
from rest_framework_simplejwt.tokens import AccessToken, RefreshToken
from rest_framework import generics, mixins
from rest_framework import viewsets
from rest_framework.pagination import LimitOffsetPagination
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import status, generics, filters
from datetime import datetime
from openpyxl import Workbook

# -*- coding: utf-8 -*-
jwt_payload_handler = api_settings.JWT_PAYLOAD_HANDLER
jwt_encode_handler = api_settings.JWT_ENCODE_HANDLER

from rest_framework_swagger.views import get_swagger_view


# schema_view = get_swagger_view(title='Pastebin API')
#
# urlpatterns = [
#     url(r'^$', schema_view)
# ]


@api_view(['POST'])
@permission_classes([AllowAny, ])
def register_user(request):
    if request.method == 'POST':
        serializer = CustomuserSerializer4(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


# @api_view(['POST'])
# @permission_classes([AllowAny, ])
# def set_password(request):
#     try:
#         old_password = request.data.get('old_password')
#         new_password = request.data.get('new_password')
#         user = request.user
#         if user.check_password(old_password):
#             user.set_password(new_password)
#             user.save()
#             result = {
#                 'status': 1,
#                 'msg': 'User password updated',
#                 'user': CustomuserSerializer(user, many=False, context={"request": request}).data
#             }
#             return Response(result, status=status.HTTP_200_OK)
#         else:
#             result = {
#                 'status': 1,
#                 'msg': 'User old password wrong'
#             }
#             return Response(result, status=status.HTTP_200_OK)
#     except KeyError:
#         res = {
#             'status': 0,
#             'msg': 'Please set all reqiured fields'
#         }
#         return Response(res)


@api_view(['POST'])
@permission_classes([AllowAny, ])
def register(request):
    try:
        username = request.data.get('username')
        last_name = request.data.get('last_name')
        phone_no = request.data.get('phone_no')
        email = request.data.get('email')
        password = request.data.get('password')
        confirm_password = request.data.get('confirm_password')
        if not login:
            res = {
                'msg': 'Login empty',
                'status': 0,
            }
            return Response(res)

        user = User.objects.filter(email=email).first()
        if not user:
            user = User.objects.create(
                username=username,
                last_name=last_name,
                email=email,
                phone_no=phone_no,
                password=password,
                complete=1
            )
        elif user:
            res = {
                'msg': 'User exits',
                'status': 2,
            }
            return Response(res)
        elif password != confirm_password:
            res = {
                'msg': 'Password not equal',
                'status': 2,
            }
            return Response(res)

        if user:
            # payload = jwt_payload_handler(user)
            # token = jwt_encode_handler(payload)
            result = {
                'status': 1,
                'user': CustomuserSerializer(user, many=False, context={"request": request}).data,
                # 'token': token
            }
            return Response(result, status=status.HTTP_200_OK)
        else:
            res = {
                'status': 0,
                'msg': 'Can not authenticate with the given credentials or the account has been deactivated'
            }
            return Response(res, status=status.HTTP_403_FORBIDDEN)
    except KeyError:
        res = {
            'status': 0,
            'msg': 'Please set all reqiured fields'
        }
        return Response(res)


@api_view(['POST'])
@permission_classes([AllowAny, ])
def userlogin(request):
    try:
        email = request.data.get('email')
        password = request.data.get('password')
        if not login:
            res = {
                'msg': 'Login empty',
                'status': 0,
            }
            return Response(res)

        user = User.objects.filter(email=email).first()
        if not user:
            res = {
                'msg': 'email or password wrond',
                'status': 403,
            }
            return Response(res, status=status.HTTP_403_FORBIDDEN)

        if user and user.check_password(password):
            if user.complete == 0:
                res = {
                    'msg': 'email sms code not check',
                    'status': 0,
                }
                return Response(res)
            token = RefreshToken.for_user(user)
            result = {

                'user': UserSerializer(user, many=False, context={"request": request}).data,
                'access': str(token.access_token),
                'refresh': str(token),

            }
            user.save()
            return Response(result, status=status.HTTP_200_OK)
        else:
            res = {
                'status': 0,
                'msg': 'Can not authenticate with the given credentials or the account has been deactivated'
            }
            return Response(res, status=status.HTTP_403_FORBIDDEN)
    except KeyError:
        res = {
            'status': 0,
            'msg': 'Please set all reqiured fields'
        }

        return Response(res)


@api_view(['POST'])
@permission_classes([AllowAny, ])
def login(request):
    try:
        email = request.data.get('email')
        password = request.data.get('password')
        firebase_token = request.data.get('firebase_token')
        android = request.data.get('android')
        if not login:
            res = {
                'msg': 'Login empty',
                'status': 0,
            }
            return Response(res)

        user = User.objects.filter(email=email).first()
        if not user:
            res = {
                'msg': 'email or password wrond',
                'status': 403,
            }
            return Response(res, status=status.HTTP_403_FORBIDDEN)

        if user and user.check_password(password):
            if user.complete == 0:
                res = {
                    'msg': 'email sms code not check',
                    'status': 0,
                }
                return Response(res)
            token = RefreshToken.for_user(user)
            result = {

                'user': UserSerializer(user, many=False, context={"request": request}).data,
                'access': str(token.access_token),
                'refresh': str(token),

            }
            if android == '1':
                user.firebase_token = firebase_token
            elif android == '0':
                user.firebase_token_front = firebase_token
            user.save()
            return Response(result, status=status.HTTP_200_OK)
        else:
            res = {
                'status': 0,
                'msg': 'Can not authenticate with the given credentials or the account has been deactivated'
            }
            return Response(res, status=status.HTTP_403_FORBIDDEN)
    except KeyError:
        res = {
            'status': 0,
            'msg': 'Please set all reqiured fields'
        }

        return Response(res)


class LogoutView(generics.GenericAPIView):
    serializer_class = LogoutSerializer
    permission_classes = (permissions.IsAuthenticated,)

    def post(self, request):
        serializer = self.serializer_class(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(status=status.HTTP_204_NO_CONTENT)


@api_view(['POST'])
@permission_classes([IsAuthenticated, ])
def set_password(request):
    try:
        old_password = request.data.get('old_password')
        new_password = request.data.get('new_password')
        user = request.user
        if user.check_password(old_password):
            user.set_password(new_password)
            user.newpassword = new_password
            user.save()
            result = {
                'status': 1,
                'msg': 'User password updated',
                'user': CustomuserSerializer(user, many=False, context={"request": request}).data
            }
            return Response(result, status=status.HTTP_200_OK)
        else:
            result = {
                'status': 1,
                'msg': 'User old password wrong'
            }
            return Response(result, status=status.HTTP_403_FORBIDDEN)
    except KeyError:
        res = {
            'status': 0,
            'msg': 'Please set all reqiured fields'
        }
        return Response(res)


@api_view(['PUT'])
@permission_classes([IsAuthenticated, ])
def profil(request):
    try:
        user = request.user
        username = request.data.get('username', user.username)
        last_name = request.data.get('last_name', user.last_name)
        patronymic_name = request.data.get('patronymic_name', user.patronymic_name)
        email = request.data.get('email', user.email)
        phone_no = request.data.get('phone_no', user.phone_no)
        image = request.data.get('image', user.image)
        unvoni = request.data.get('unvoni', user.unvoni)
        xonasi = request.data.get('xonasi', user.xonasi)
        firebase_token = request.data.get('firebase_token', user.firebase_token)

        user.username = username
        user.last_name = last_name
        user.patronymic_name = patronymic_name
        user.email = email
        user.phone_no = phone_no
        user.image = image
        user.xonasi = xonasi
        user.unvoni = unvoni
        user.firebase_token = firebase_token
        user.save()

        result = {
            'status': 1,
            'msg': 'User updated',
            'user': CustomuserSerializer3(user, many=False, context={"request": request}).data
        }
        return Response(result, status=status.HTTP_200_OK)
    except KeyError:
        res = {
            'status': 0,
            'msg': 'Please set all reqiured fields'
        }
        return Response(res)


@api_view(['PUT'])
@permission_classes([IsAuthenticated, ])
def updateFirebase_token(request):
    try:
        user = request.user

        firebase_token = request.data.get('firebase_token', user.firebase_token)

        user.firebase_token = firebase_token
        user.save()

        result = {
            'status': 1,
            'msg': 'Token updated',
            'user': CustomuserSerializer3(user, many=False, context={"request": request}).data
        }
        return Response(result, status=status.HTTP_200_OK)
    except KeyError:
        res = {
            'status': 0,
            'msg': 'Please set all reqiured fields'
        }
        return Response(res)


@api_view(['GET'])
@permission_classes([IsAuthenticated, ])
def me(request):
    try:
        user = request.user
        result = {
            'status': 1,
            'user': ProfilSerializerMe(user, many=False, context={"request": request}).data
        }
        return Response(result, status=status.HTTP_200_OK)
    except KeyError:
        res = {
            'status': 0,
            'msg': 'Please set all reqiured fields'
        }
        return Response(res)


# class ProfileAPI(APIView):
#     def get(self, request, *args, **kwargs):
#         user = get_object_or_404(User, pk=kwargs['user_id'])
#         profile_serializer = ProfilSerializer(user.profile)
#         return Response(profile_serializer.data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def getProfile(request):
    user = request.user
    serializer = ProfilSerializerMe(user, many=False)
    return Response(serializer.data)


@permission_classes([IsAuthenticated, ])
class UserDetailView(generics.RetrieveAPIView):
    queryset = User.objects.all()
    serializer_class = CustomuserSerializer3


# @permission_classes([IsAuthenticated, ])
# class MessageDetailView(generics.RetrieveAPIView):
#     queryset = Message.objects.all()
#     serializer_class = GetMessageSerializerAll2
#     filterset_class = MessageFilter1
#     filter_backends = [DjangoFilterBackend, filters.SearchFilter]
#
#     def get(self, request, *args, **kwargs):
#
#         # queryset = Message.objects.filter(user_id=self.request.user.id, id=kwargs['pk'])
#         queryset = Message.objects.filter(user_id=self.request.user.id)
#
#         for x in queryset:
#             if x.status != 'bajarilmadi' and x.status != 'bajarildi':
#                 queryset.update(status="qabulqildi")
#         serializer = GetMessageSerializerAll2(queryset, many=True, context={"request": request})
#         return Response(serializer.data, status=status.HTTP_200_OK)


class MessageDetailView(generics.ListAPIView, mixins.ListModelMixin, viewsets.GenericViewSet):
    serializer_class = GetMessageSerializerAll2
    permission_classes = [IsAuthenticated]
    queryset = Message.objects.order_by('-id').all()
    # pagination_class = LimitOffsetPagination
    filterset_class = MessageFilter1
    pagination_class = LargeResultsSetPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]

    def get_queryset(self):
        queryset = Message.objects.filter(user_id=self.request.user.id, id=self.request.GET['id'])

        for x in queryset:
            if x.status != 'bajarilmadi' and x.status != 'bajarildi':
                queryset.update(status="qabulqildi")
        return queryset


class MessageDetailView2(generics.ListAPIView, mixins.ListModelMixin, viewsets.GenericViewSet):
    serializer_class = GetMessageSerializerAll2
    permission_classes = [IsAuthenticated]
    queryset = Message.objects.order_by('-id').all()
    filterset_class = MessageFilter1
    pagination_class = LargeResultsSetPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]

    def get_queryset(self):
        queryset = Message.objects.filter(created_user=self.request.user.id, id=self.request.GET['id'])
        event = Message.objects.first()
        for x in queryset:
            if x.status2 == 'kurilmagan' and x.status == 'bajarildi':
                queryset.update(status2="kurildi")
        if event.state:
            return queryset


class MessageDetailPostViewID(generics.ListAPIView, mixins.ListModelMixin, viewsets.GenericViewSet):
    serializer_class = GetMessageSerializerAll2
    permission_classes = [IsAuthenticated]
    queryset = Message.objects.order_by('-id').all()
    pagination_class = LimitOffsetPagination
    filterset_class = MessageFilter1
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]

    def post(self, request, *args, **kwargs):
        queryset = Message.objects.filter(user_id=self.request.user.id, id=self.request.GET['id'])
        message = Message.objects.get(user_id=self.request.user.id, id=self.request.GET['id'])
        for x in queryset:
            text_employee = request.data['text_employee']
            if x.status != 'bajarilmadi' and x.status != 'kechikibbajarildi':
                queryset.update(status="bajarildi", text_employee=text_employee)
            elif x.status == 'bajarilmadi':
                queryset.update(status="kechikibbajarildi", text_employee=text_employee)

        files = request.FILES.getlist('file')
        for file in files:
            filesave = File_employee.objects.create(
                file=file,
                message=message
            ).save()
        serializer = GetMessageSerializerAll2(queryset, many=True, context={"request": request})
        return Response(serializer.data, status=status.HTTP_200_OK)


@api_view(['POST'])
@permission_classes([IsAdminUser, ])
def post_message(request):
    try:
        item_serializer = UserArraySerializer(data=request.data)
        item_serializer.is_valid(raise_exception=True)
        user = item_serializer.data['user']
        # user = request.data["user"]
        text = request.data['text']
        end_time = request.data['end_time']
        for x in user:
            user1 = User.objects.filter(id=x)
            message = Message.objects.create(
                # user=request.user,
                user_id=x,
                text=text,
                end_time=end_time,
                status='yuborildi',
                created_user=request.user.id,

            )
            message.save()
            files = request.FILES.getlist('file')
            for file in files:
                File.objects.create(
                    file=file,
                    message=message
                )
            # funksiya
            if user1.last().firebase_token and user1.last().firebase_token != 'null':
                send_firebase_message(user1.last().firebase_token, 'Hujjat Almashinuv Tizimi', 'Yangi xabar')
            if user1.last().firebase_token_front and user1.last().firebase_token_front != 'null':
                send_firebase_message1(user1.last().firebase_token_front, 'Hujjat Almashinuv Tizimi1', 'Yangi xabar1')
            else:
                pass

        result = {
            'status': 1,
            'msg': 'add_message',
            'message': PostMessageSerializer(message, many=False, context={"request": request}).data,
        }
        return Response(result, status=status.HTTP_200_OK)

    except KeyError:
        res = {
            'status': 0,
            'msg': 'erorr add_message'
        }
        return Response(res)


@api_view(['PUT'])
@permission_classes([IsAuthenticated, ])
def put_message(request):
    try:
        user = request.user
        username = request.data.get('username', user.username)
        last_name = request.data.get('last_name', user.last_name)
        patronymic_name = request.data.get('patronymic_name', user.patronymic_name)
        email = request.data.get('email', user.email)
        phone_no = request.data.get('phone_no', user.phone_no)
        image = request.data.get('image', user.image)
        unvoni = request.data.get('unvoni', user.unvoni)
        xonasi = request.data.get('xonasi', user.xonasi)
        firebase_token = request.data.get('firebase_token', user.firebase_token)

        user.username = username
        user.last_name = last_name
        user.patronymic_name = patronymic_name
        user.email = email
        user.phone_no = phone_no
        user.image = image
        user.xonasi = xonasi
        user.unvoni = unvoni
        user.firebase_token = firebase_token
        user.save()

        result = {
            'status': 1,
            'msg': 'User updated',
            'user': CustomuserSerializer3(user, many=False, context={"request": request}).data
        }
        return Response(result, status=status.HTTP_200_OK)
    except KeyError:
        res = {
            'status': 0,
            'msg': 'Please set all reqiured fields'
        }
        return Response(res)


# class MessageUpdateView(generics.UpdateAPIView, generics.DestroyAPIView):
#     queryset = Message.objects.all()
#     serializer_class = PutMessageSerializer
#
#     def get_queryset(self):
#         queryset = Message.objects.filter(created_user=self.request.user.id)
#         return queryset


class MessageUpdateView(APIView):
    def get_object(self, pk):
        try:
            return Message.objects.get(pk=pk, created_user=self.request.user.id, status='yuborildi')
        except Message.DoesNotExist:
            raise Http404

    def put(self, request, pk=None, format=None):

        message = Message.objects.get(pk=pk, created_user=self.request.user.id, status='yuborildi')
        serializer = PutMessageSerializer(instance=message, data=request.data, partial=True,
                                          context={'request': request})
        serializer.is_valid(raise_exception=True)
        serializer.save()
        files = request.FILES.getlist('file')
        if request.FILES.getlist('file'):
            file = File.objects.filter(message=message)
            file.delete()
        for file in files:
            File.objects.filter(message=message).create(
                file=file,
                message=message
            )
        response = Response()

        response.data = {
            'message': 'Message Updated Successfully',
            'data': serializer.data,
        }

        return response

    def delete(self, request, pk, format=None):
        message = Message.objects.get(pk=pk, created_user=self.request.user.id, status='yuborildi')

        message.delete()

        return Response({
            'message': 'Message Deleted Successfully'
        })


class MalumotuchunUpdateView(APIView):
    def get_object(self, pk):
        try:
            return MalumotUchun.objects.get(pk=pk, created_user=self.request.user.id, status='kurilmagan')
        except MalumotUchun.DoesNotExist:
            raise Http404

    def put(self, request, pk=None, format=None):

        malumotuchun = MalumotUchun.objects.get(pk=pk, created_user=self.request.user.id, status='kurilmagan')
        serializer = PostMalumotUchunSerializer(instance=malumotuchun, data=request.data, partial=True,
                                          context={'request': request})
        serializer.is_valid(raise_exception=True)
        serializer.save()
        files = request.FILES.getlist('file')
        if request.FILES.getlist('file'):
            file = File.objects.filter(malumotuchun=malumotuchun)
            file.delete()
        for file in files:
            File.objects.filter(malumotuchun=malumotuchun).create(
                file=file,
                malumotuchun=malumotuchun
            )
        response = Response()

        response.data = {
            'message': 'Message Updated Successfully',
            'data': serializer.data,
        }

        return response

    def delete(self, request, pk, format=None):
        malumotuchun = MalumotUchun.objects.get(pk=pk, created_user=self.request.user.id, status='kurilmagan')

        malumotuchun.delete()

        return Response({
            'message': 'Message Deleted Successfully'
        })


@api_view(['POST'])
@permission_classes([IsAdminUser, ])
def post_malumotuchun(request):
    try:
        item_serializer = UserArraySerializer(data=request.data)
        item_serializer.is_valid(raise_exception=True)
        user = item_serializer.data['user']
        text = request.data['text']
        for x in user:
            user1 = User.objects.filter(id=x)
            message = MalumotUchun.objects.create(
                user_id=x,
                text=text,
                created_user=request.user.id,

            )
            message.save()
            # funksiya
            if user1.last().firebase_token and user1.last().firebase_token != 'null':
                send_firebase_message(user1.last().firebase_token, 'Hujjat Almashinuv Tizimi', 'Yangi xabar')
            if user1.last().firebase_token_front and user1.last().firebase_token_front != 'null':
                send_firebase_message1(user1.last().firebase_token_front, 'Hujjat Almashinuv Tizimi1', 'Yangi xabar1')
            else:
                pass
            files = request.FILES.getlist('file')
            for file in files:
                File.objects.create(
                    file=file,
                    malumotuchun=message
                )
        result = {
            'status': 1,
            'msg': 'add_message',
            'message': PostMalumotUchunSerializer(message, many=False, context={"request": request}).data,
        }
        return Response(result, status=status.HTTP_200_OK)

    except KeyError:
        res = {
            'status': 0,
            'msg': 'erorr add_message'
        }
        return Response(res)


class MalumotDetailView(generics.ListAPIView, mixins.ListModelMixin, viewsets.GenericViewSet):
    serializer_class = MalumotSerializerAll
    permission_classes = [IsAuthenticated]
    queryset = MalumotUchun.objects.order_by('-id').all()
    # pagination_class = LimitOffsetPagination
    filterset_class = MalumotUchunFilter
    pagination_class = LargeResultsSetPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]

    def get_queryset(self):
        queryset = MalumotUchun.objects.filter(user_id=self.request.user.id)
        return queryset


class MalumotDetailView2(generics.ListAPIView, mixins.ListModelMixin, viewsets.GenericViewSet):
    serializer_class = MalumotSerializerAll
    permission_classes = [IsAuthenticated]
    queryset = Message.objects.order_by('-id').all()
    filterset_class = MalumotUchunFilter
    pagination_class = LargeResultsSetPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]

    def get_queryset(self):
        queryset = MalumotUchun.objects.filter(user_id=self.request.user.id, id=self.request.GET['id'])

        for x in queryset:
            if x.status == 'kurilmagan':
                queryset.update(status="kurildi")
            return queryset


class MessageViewSet(generics.ListAPIView, mixins.ListModelMixin, viewsets.GenericViewSet):
    serializer_class = GetMessageSerializerAll2
    permission_classes = [IsAuthenticated]
    queryset = Message.objects.order_by('-id').all()
    # pagination_class = LimitOffsetPagination
    filterset_class = MessageFilter
    pagination_class = LargeResultsSetPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]

    def get_queryset(self):
        queryset = Message.objects.filter(created_user=self.request.user.id)
        event = Message.objects.first()
        if event.state:
            return queryset


class MessageDetailView2(generics.ListAPIView, mixins.ListModelMixin, viewsets.GenericViewSet):
    serializer_class = GetMessageSerializerAll2
    permission_classes = [IsAuthenticated]
    queryset = Message.objects.order_by('-id').all()
    filterset_class = MessageFilter1
    pagination_class = LargeResultsSetPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]

    def get_queryset(self):
        queryset = Message.objects.filter(created_user=self.request.user.id, id=self.request.GET['id'])
        event = Message.objects.first()
        for x in queryset:
            if x.status2 == 'kurilmagan' and x.status == 'bajarildi':
                queryset.update(status2="kurildi")
        if event.state:
            return queryset


class MalumotuchunViewSet(generics.ListAPIView, mixins.ListModelMixin, viewsets.GenericViewSet):
    serializer_class = MalumotSerializerAll
    permission_classes = [IsAuthenticated]
    queryset = MalumotUchun.objects.order_by('-id').all()
    # pagination_class = LimitOffsetPagination
    filterset_class = MalumotUchunFilter
    pagination_class = LargeResultsSetPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]

    def get_queryset(self):
        queryset = MalumotUchun.objects.filter(created_user=self.request.user.id)
        return queryset


class GetUserMessageViewSet(generics.ListAPIView, mixins.ListModelMixin, viewsets.GenericViewSet):
    serializer_class = GetMessageSerializerAll2
    permission_classes = [IsAuthenticated]
    queryset = Message.objects.order_by('-id').all()
    # pagination_class = LimitOffsetPagination
    filterset_class = MessageFilter
    pagination_class = LargeResultsSetPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]

    def get_queryset(self):
        queryset = Message.objects.filter(user_id=self.request.user.id)
        return queryset


#     if self.request.user.is_staff == False:
#         queryset = Message.objects.filter(user_id=self.request.user.id)
#         return queryset
#     else:
#         queryset = Message.objects.filter(created_user=self.request.user.id)
#         event = Message.objects.first()
#         if event.state:
#             return queryset


class GetAllMessageViewSet(generics.ListAPIView, mixins.ListModelMixin, viewsets.GenericViewSet):
    serializer_class = GetAllMessageSerializer
    permission_classes = [IsAuthenticated]
    queryset = AllNotification.objects.order_by('-id').all()
    pagination_class = LimitOffsetPagination
    # filterset_class = MessageFilter
    # pagination_class = LargeResultsSetPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]


from django.db.models import Q


class GetUsersViewSet(generics.ListAPIView, mixins.ListModelMixin, viewsets.GenericViewSet):
    serializer_class = GetUsersSerializer
    permission_classes = [IsAdminUser]
    queryset = User.objects.order_by('-id').all()
    filterset_class = UserFilter
    pagination_class = LargeResultsSetPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]

    def get_queryset(self):
        queryset = User.objects.filter(~Q(id=self.request.user.id), ~Q(id=1))
        return queryset


class IsSuperUser(IsAdminUser):
    def has_permission(self, request, view):
        return bool(request.user and request.user.superuser)


class GetUsersStatisticsViewSet(generics.ListAPIView, mixins.ListModelMixin, viewsets.GenericViewSet):
    serializer_class = GetUsersStatisticsSerializer
    permission_classes = [IsSuperUser]
    queryset = User.objects.order_by('-id').all()
    filterset_class = UserFilter
    pagination_class = LargeResultsSetPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]


from django.db.models import Q, Count


class StatisticsAll(APIView):
    permission_classes = [IsSuperUser]  # policy attribute

    def get(self, request):
        yuborildi = Message.objects.filter(status='yuborildi').count()
        qabulqildi = Message.objects.filter(status='qabulqildi').count()
        bajarildi = Message.objects.filter(status='bajarildi').count()
        kechikibbajarildi = Message.objects.filter(status='kechikibbajarildi').count()
        bajarilmadi = Message.objects.filter(status='bajarilmadi').count()
        content = {
            'yuborildi': yuborildi,
            'qabulqildi': qabulqildi,
            'bajarildi': bajarildi,
            'kechikibbajarildi': kechikibbajarildi,
            'bajarilmadi': bajarilmadi,
        }

        return Response(content)


class StatisticsSolo(APIView):
    # permission_classes = [IsSuperUser]  # policy attribute

    def get(self, request):
        user = User.objects.get(id=self.request.user.id)
        yuborildi = Message.objects.filter(status='yuborildi', user_id=user.id).count()
        qabulqildi = Message.objects.filter(status='qabulqildi', user_id=user.id).count()
        bajarildi = Message.objects.filter(status='bajarildi', user_id=user.id).count()
        kechikibbajarildi = Message.objects.filter(status='kechikibbajarildi', user_id=user.id).count()
        bajarilmadi = Message.objects.filter(status='bajarilmadi', user_id=user.id).count()
        content = {
            'yuborildi': yuborildi,
            'qabulqildi': qabulqildi,
            'bajarildi': bajarildi,
            'kechikibbajarildi': kechikibbajarildi,
            'bajarilmadi': bajarilmadi,
        }

        return Response(content)


class StatisticsSolo2(APIView):
    # permission_classes = [IsSuperUser]  # policy attribute

    def get(self, request):
        user = User.objects.get(id=self.request.user.id)
        yuborildi = Message.objects.filter(status='yuborildi', created_user=user.id).count()
        qabulqildi = Message.objects.filter(status='qabulqildi', created_user=user.id).count()
        bajarildi = Message.objects.filter(status='bajarildi', created_user=user.id).count()
        kechikibbajarildi = Message.objects.filter(status='kechikibbajarildi', created_user=user.id).count()
        bajarilmadi = Message.objects.filter(status='bajarilmadi', created_user=user.id).count()
        content = {
            'yuborildi': yuborildi,
            'qabulqildi': qabulqildi,
            'bajarildi': bajarildi,
            'kechikibbajarildi': kechikibbajarildi,
            'bajarilmadi': bajarilmadi,
        }

        return Response(content)


class UsersStatisticsViewSet(generics.ListAPIView, mixins.ListModelMixin, viewsets.GenericViewSet):
    serializer_class = UsersStatisticsSerializer
    permission_classes = [IsAuthenticated]
    queryset = User.objects.order_by('-id').all()
    filterset_class = UserFilter
    pagination_class = LargeResultsSetPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]


class UsersStatisticsViewSet2(generics.ListAPIView, mixins.ListModelMixin, viewsets.GenericViewSet):
    serializer_class = UsersStatisticsSerializer2
    permission_classes = [IsAuthenticated]
    queryset = User.objects.order_by('-id').all()
    filterset_class = UserFilter
    pagination_class = LargeResultsSetPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]


class GetKorxonaViewSet(generics.ListAPIView, mixins.ListModelMixin, viewsets.GenericViewSet):
    serializer_class = KorxonaSerializer
    permission_classes = [AllowAny]
    queryset = Korxona.objects.order_by('-id').all()
    # filterset_class = MessageFilter
    pagination_class = LargeResultsSetPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]


from django.shortcuts import render
from django.http import HttpResponse, JsonResponse, Http404

# Create your views here.
import csv


@api_view(['GET'])
@permission_classes([IsAuthenticated, ])
def export_movies_to_xlsx(request):
    user_queryset = User.objects.all()

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-movies.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'User'

    # Define the titles for columns
    columns = [
        'Raqami',
        'Ismi',
        'Familyasi',
        'Sharifi',
        'yuborildi',
        'qabulqildi',
        'bajarildi',
        'kechikibbajarildi',
        'bajarilmadi',

    ]
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    nomer = 0
    for s in user_queryset:
        row_num += 1
        nomer += 1
        yuborildi = Message.objects.filter(user=s, status='yuborildi').count()
        qabulqildi = Message.objects.filter(user=s, status='qabulqildi').count()
        bajarildi = Message.objects.filter(user=s, status='bajarildi').count()
        kechikibbajarildi = Message.objects.filter(user=s, status='kechikibbajarildi').count()
        bajarilmadi = Message.objects.filter(user=s, status='bajarilmadi').count()

        # Define the data for each cell in the row
        row = [
            nomer,
            s.username,
            s.last_name,
            s.patronymic_name,
            yuborildi,
            qabulqildi,
            bajarildi,
            kechikibbajarildi,
            bajarilmadi

        ]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated, ])
def solo_export_movies_to_xlsx(request):
    user_queryset = User.objects.all()

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={name}-{date}.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d', ),
        name=request.user.username
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'User'

    # Define the titles for columns
    columns = [
        'Raqami',
        'Ismi',
        'Familyasi',
        'Sharifi',
        'yuborildi',
        'qabulqildi',
        'bajarildi',
        'kechikibbajarildi',
        'bajarilmadi',

    ]
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    nomer = 0
    for s in user_queryset:
        row_num += 1
        nomer += 1
        yuborildi = Message.objects.filter(user=s, created_user=request.user.id, status='yuborildi').count()
        qabulqildi = Message.objects.filter(user=s, created_user=request.user.id, status='qabulqildi').count()
        bajarildi = Message.objects.filter(user=s, created_user=request.user.id, status='bajarildi').count()
        kechikibbajarildi = Message.objects.filter(user=s, created_user=request.user.id,
                                                   status='kechikibbajarildi').count()
        bajarilmadi = Message.objects.filter(user=s, created_user=request.user.id, status='bajarilmadi').count()

        # Define the data for each cell in the row
        row = [
            nomer,
            s.username,
            s.last_name,
            s.patronymic_name,
            yuborildi,
            qabulqildi,
            bajarildi,
            kechikibbajarildi,
            bajarilmadi

        ]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated, ])
def solo_export_movies_to_xlsx2(request):
    user_queryset = User.objects.all()

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={name}-{date}.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d', ),
        name=request.user.username
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'User'

    # Define the titles for columns
    columns = [
        'Raqami',
        'Ismi',
        'Familyasi',
        'Sharifi',
        'yuborildi',
        'qabulqildi',
        'bajarildi',
        'kechikibbajarildi',
        'bajarilmadi',

    ]
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    nomer = 0
    for s in user_queryset:
        row_num += 1
        nomer += 1
        yuborildi = Message.objects.filter(user=request.user.id, created_user=s.id, status='yuborildi').count()
        qabulqildi = Message.objects.filter(user=request.user.id, created_user=s.id, status='qabulqildi').count()
        bajarildi = Message.objects.filter(user=request.user.id, created_user=s.id, status='bajarildi').count()
        kechikibbajarildi = Message.objects.filter(user=request.user.id, created_user=s.id,
                                                   status='kechikibbajarildi').count()
        bajarilmadi = Message.objects.filter(user=request.user.id, created_user=s.id, status='bajarilmadi').count()

        # Define the data for each cell in the row
        row = [
            nomer,
            s.username,
            s.last_name,
            s.patronymic_name,
            yuborildi,
            qabulqildi,
            bajarildi,
            kechikibbajarildi,
            bajarilmadi

        ]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response
